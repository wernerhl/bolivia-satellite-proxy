"""Parse INE/Aduana annual imports xlsx into monthly CIF totals.

Source: data/official/aduana_raw/{2024,2025,2026}_imports.xlsx — detail-
record exports from the Aduana Nacional via INE. Each row is an
individual import declaration with columns including date, gross
weight, FOB/CIF value, etc.

We aggregate to monthly CIF USD and monthly gross weight KG.
"""
from __future__ import annotations

import re
import sys
from pathlib import Path

import pandas as pd

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
from _common import abs_path, ensure_dir, load_env  # noqa: E402

load_env()


RAW_DIR = abs_path("data/official/aduana_raw")
OUT = abs_path("data/official/aduana_imports.csv")


def _find_date_col(columns: list[str]) -> str | None:
    pats = ["FECHA", "PERIODO", "MES", "PERIOD"]
    for c in columns:
        cu = str(c).upper()
        if any(p in cu for p in pats):
            return c
    return None


def _find_value_col(columns: list[str], hint: str) -> str | None:
    for c in columns:
        cu = str(c).upper()
        if hint in cu and ("VALOR" in cu or "USD" in cu or "CIF" in cu or "FOB" in cu):
            return c
    # Fallback: just the hint
    for c in columns:
        if hint in str(c).upper():
            return c
    return None


def parse_year(path: Path, year: int) -> pd.DataFrame:
    """Stream through the sheet with openpyxl (memory-safe on 374MB sheets).

    INE/Aduana customs-microdata schema (verified on 2024 file):
      col 0  GESTION   year
      col 1  MES       month (1-12 integer)
      col 24 KILOS     gross weight
      col 26 FOB       FOB value in USD
      col 27 ADU       ADUANA (CIF-at-customs) value in USD
    """
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows_iter = ws.iter_rows(values_only=True)
    _ = next(rows_iter)  # header

    GESTION_C, MES_C, KILOS_C, FOB_C, ADU_C = 0, 1, 24, 26, 27
    by_month: dict[int, dict[str, float]] = {}

    for row in rows_iter:
        if row is None or row[MES_C] is None:
            continue
        try:
            month = int(row[MES_C])
        except Exception:
            continue
        if not 1 <= month <= 12:
            continue
        agg = by_month.setdefault(month, {
            "imports_usd_cif": 0.0, "imports_fob_usd": 0.0,
            "imports_kg": 0.0, "n_records": 0,
        })
        for (idx, key) in ((ADU_C, "imports_usd_cif"),
                           (FOB_C, "imports_fob_usd"),
                           (KILOS_C, "imports_kg")):
            v = row[idx]
            if v is None:
                continue
            try:
                agg[key] += float(v)
            except Exception:
                pass
        agg["n_records"] += 1

    records = []
    for m, stats in sorted(by_month.items()):
        records.append({
            "date": pd.Timestamp(year=year, month=m, day=1),
            "imports_usd_cif": stats["imports_usd_cif"],
            "imports_fob_usd": stats["imports_fob_usd"],
            "imports_kg": stats["imports_kg"],
            "n_records": stats["n_records"],
        })
    return pd.DataFrame(records)


def main() -> None:
    frames = []
    for fp in sorted(RAW_DIR.glob("*_imports.xlsx")):
        y_match = re.match(r"(\d{4})", fp.name)
        if not y_match:
            continue
        year = int(y_match.group(1))
        print(f"[..] parsing {fp.name}")
        df = parse_year(fp, year)
        if not df.empty:
            frames.append(df)
            print(f"[ok] {fp.name}: {len(df)} months, "
                  f"range {df['imports_usd_cif'].min():.0f} .. "
                  f"{df['imports_usd_cif'].max():.0f} USD")
    if frames:
        all_df = pd.concat(frames, ignore_index=True).sort_values("date")
        ensure_dir(OUT.parent)
        all_df.to_csv(OUT, index=False)
        print(f"[ok] wrote {OUT} ({len(all_df)} rows)")


if __name__ == "__main__":
    main()
